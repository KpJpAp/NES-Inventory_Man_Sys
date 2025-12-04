import tkinter as tk
from PIL import Image, ImageTk
import tkinter.font as tkfont
from tkinter import messagebox, ttk, filedialog
import sqlite3
import openpyxl
import time
from openpyxl.styles import Font
from datetime import datetime
import sys
import os
from threading import Timer, Thread

# Modern UI Colors
COLORS = {
    'bg': '#f8fafc',
    'card': '#ffffff',
    'primary': '#3b82f6',
    'primary_hover': '#2563eb',
    'secondary': '#64748b',
    'accent': '#10b981',
    'danger': '#ef4444',
    'danger_hover': '#dc2626',
    'text': '#1e293b',
    'text_light': '#64748b',
    'border': '#e2e8f0',
    'success': '#22c55e'
}

# --- Database Connection Management ---
def get_db_connection(max_retries=3):
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'inventory.db')
    print(f"Attempting to connect to database at: {db_path}")  # Debug info

    for attempt in range(max_retries):
        try:
            connection = sqlite3.connect(db_path)
            # Test the connection
            cursor = connection.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='inventory'")
            if not cursor.fetchone():
                # Table doesn't exist, create it
                cursor.execute('''
                CREATE TABLE IF NOT EXISTS inventory (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT NOT NULL CHECK(length(item_name) <= 100),
                    quantity INTEGER NOT NULL CHECK(quantity >= 0),
                    price REAL NOT NULL CHECK(price >= 0),
                    updated_by TEXT NOT NULL CHECK(length(updated_by) <= 50),
                    low_stock_threshold INTEGER DEFAULT 10 CHECK(low_stock_threshold >= 0)
                )
                ''')
                connection.commit()
            return connection
        except sqlite3.Error as e:
            if attempt == max_retries - 1:  # Last attempt
                messagebox.showerror("Database Error", f"Failed to connect to database after {max_retries} attempts: {e}\nPath: {db_path}")
                raise
            time.sleep(1)  # Wait before retrying

def safe_commit(connection):
    try:
        connection.commit()
    except sqlite3.Error as e:
        connection.rollback()
        raise sqlite3.Error(f"Failed to commit transaction: {e}")

def initialize_database():
    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Create tables if they don't exist
        c.execute('''
        CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_name TEXT NOT NULL CHECK(length(item_name) <= 100),
            quantity INTEGER NOT NULL CHECK(quantity >= 0),
            price REAL NOT NULL CHECK(price >= 0),
            updated_by TEXT NOT NULL CHECK(length(updated_by) <= 50),
            low_stock_threshold INTEGER DEFAULT 10 CHECK(low_stock_threshold >= 0)
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            item_id INTEGER,
            item_name TEXT,
            user TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Create indexes if they don't exist
        c.execute("CREATE INDEX IF NOT EXISTS idx_item_name ON inventory(item_name)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_quantity ON inventory(quantity)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_price ON inventory(price)")

        conn.commit()
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Failed to initialize database: {e}")
        exit(1)

# --- GUI Setup and Functions ---
def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')

class RoundedFrame(tk.Frame):
    def __init__(self, parent, bg_color=COLORS['card'], border_color=COLORS['border'],
                 corner_radius=12, border_width=1, **kwargs):
        super().__init__(parent, bg=bg_color, highlightthickness=0, **kwargs)
        self.bg_color = bg_color
        self.border_color = border_color
        self.corner_radius = corner_radius
        self.border_width = border_width

        self.canvas = tk.Canvas(self, highlightthickness=0, bg=parent.cget('bg'))
        self.canvas.pack(fill='both', expand=True)
        self.update_idletasks()
        self.after(1, self.draw_rounded_rect)

    def draw_rounded_rect(self):
        self.canvas.delete("bg")
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()

        if w > 1 and h > 1:
            self.canvas.create_rounded_rect(0, 0, w, h, self.corner_radius,
                                          fill=self.bg_color, outline=self.border_color,
                                          width=self.border_width, tags="bg")

def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
    points = []
    for x, y in [(x1, y1 + radius), (x1, y1), (x1 + radius, y1),
                 (x2 - radius, y1), (x2, y1), (x2, y1 + radius),
                 (x2, y2 - radius), (x2, y2), (x2 - radius, y2),
                 (x1 + radius, y2), (x1, y2), (x1, y2 - radius)]:
        points.extend([x, y])
    return self.create_polygon(points, smooth=True, **kwargs)

tk.Canvas.create_rounded_rect = create_rounded_rect

def validate_non_negative_int(P):
    if P == "":
        return True
    if P.isdigit() and int(P) >= 0:
        return True
    return False

def validate_non_negative_float(P):
    if P == "":
        return True
    try:
        value = float(P)
        return value >= 0
    except ValueError:
        return False

def update_status(message):
    if 'status_var' in globals() and status_var is not None:
        status_var.set(message)

def clear_inputs():
    global selected_item_id
    selected_item_id = None
    entry_widgets["Item Name"].delete(0, tk.END)
    entry_widgets["Quantity"].delete(0, tk.END)
    entry_widgets["Price"].delete(0, tk.END)
    entry_widgets["Updated By"].delete(0, tk.END)
    add_button.config(text="Add Item", bg=COLORS['primary'])
    update_status("Ready to add new item.")

def display_inventory(filter_text="", page=1, page_size=50):
    try:
        # Clear existing items
        for row in inventory_tree.get_children():
            inventory_tree.delete(row)

        # Calculate offset for pagination
        offset = (page - 1) * page_size

        # Query with pagination
        query = """
        SELECT id, item_name, quantity, price, updated_by FROM inventory
        """
        params = ()

        if filter_text:
            query += " WHERE item_name LIKE ?"
            params = ('%' + filter_text + '%',)

        query += f" LIMIT {page_size} OFFSET {offset}"

        c.execute(query, params)

        # Insert items into tree
        for row in c.fetchall():
            inventory_tree.insert('', 'end', iid=row[0],
                                values=(row[0], row[1], row[2], f"${row[3]:.2f}",
                                       row[4] if row[4] else 'N/A'))

    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Failed to load inventory: {e}")

def load_selected_item(event):
    global selected_item_id
    selected = inventory_tree.selection()
    if selected:
        selected_item_id = int(selected[0])
        try:
            c.execute("SELECT id, item_name, quantity, price, updated_by FROM inventory WHERE id=?", (selected_item_id,))
            row = c.fetchone()
            if row:
                entry_widgets["Item Name"].delete(0, tk.END)
                entry_widgets["Item Name"].insert(0, row[1])
                entry_widgets["Quantity"].delete(0, tk.END)
                entry_widgets["Quantity"].insert(0, row[2])
                entry_widgets["Price"].delete(0, tk.END)
                entry_widgets["Price"].insert(0, row[3])
                entry_widgets["Updated By"].delete(0, tk.END)
                entry_widgets["Updated By"].insert(0, row[4] if row[4] else '')
                add_button.config(text="Update Item", bg=COLORS['accent'])
                update_status(f"Loaded item ID {selected_item_id} for editing.")
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to load item: {e}")
    else:
        clear_inputs()

def is_duplicate_name(name, exclude_id=None):
    try:
        if exclude_id is None:
            c.execute("SELECT COUNT(*) FROM inventory WHERE LOWER(item_name) = LOWER(?)", (name,))
        else:
            c.execute("SELECT COUNT(*) FROM inventory WHERE LOWER(item_name) = LOWER(?) AND id != ?", (name, exclude_id))
        count = c.fetchone()[0]
        return count > 0
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Failed to check duplicate: {e}")
        return False

def add_item():
    global selected_item_id
    name = entry_widgets["Item Name"].get().strip()
    quantity = entry_widgets["Quantity"].get().strip()
    price = entry_widgets["Price"].get().strip()
    user = entry_widgets["Updated By"].get().strip()

    if len(name) > 100:
        update_status("Item name must be 100 characters or less!")
        return
    if len(user) > 50:
        update_status("Updated By must be 50 characters or less!")
        return

    if not (name and quantity and price and user):
        update_status("All fields are required!")
        return

    name = ''.join(c for c in name if c.isprintable())
    user = ''.join(c for c in user if c.isprintable())

    try:
        quantity_int = int(quantity)
        price_float = float(price)

        if quantity_int < 0 or price_float < 0:
            update_status("Quantity and Price must be non-negative.")
            return
    except ValueError:
        update_status("Quantity must be an integer, and Price must be a number.")
        return

    if is_duplicate_name(name, exclude_id=selected_item_id):
        update_status(f"Item name '{name}' already exists.")
        return

    try:
        if selected_item_id is None:
            c.execute(
                "INSERT INTO inventory (item_name, quantity, price, updated_by) VALUES (?, ?, ?, ?)",
                (name, quantity_int, price_float, user)
            )
            item_id = c.lastrowid
            action = "Added"
        else:
            c.execute(
                "UPDATE inventory SET item_name=?, quantity=?, price=?, updated_by=? WHERE id=?",
                (name, quantity_int, price_float, user, selected_item_id)
            )
            item_id = selected_item_id
            action = "Updated"

        c.execute(
            "INSERT INTO audit_log (action, item_id, item_name, user) VALUES (?, ?, ?, ?)",
            (action, item_id, name, user)
        )

        try:
            safe_commit(conn)
            messagebox.showinfo("Success", f"Item {action.lower()} successfully!")
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to save changes: {e}")
            return

        clear_inputs()
        display_inventory()
        update_status(f"Item {action.lower()} successfully.")

    except sqlite3.Error as e:
        messagebox.showerror("Database Error", str(e))

def delete_item():
    selected = inventory_tree.selection()
    if selected:
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete the selected item?"):
            update_status("Delete cancelled.")
            return
        item_id = int(selected[0])
        try:
            c.execute("SELECT item_name, updated_by FROM inventory WHERE id=?", (item_id,))
            row = c.fetchone()
            if row:
                item_name, user = row
            else:
                item_name, user = "Unknown", "Unknown"

            c.execute("DELETE FROM inventory WHERE id=?", (item_id,))
            c.execute("INSERT INTO audit_log (action, item_id, item_name, user) VALUES (?, ?, ?, ?)",
                      ("Deleted", item_id, item_name, user))
            conn.commit()
            messagebox.showinfo("Success", "Item deleted successfully!")
            display_inventory()
            clear_inputs()
            update_status("Item deleted successfully.")
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to delete item: {e}")
    else:
        update_status("Select an item to delete.")

def show_audit_log():
    audit_window = tk.Toplevel(root)
    audit_window.title("Audit Log")
    audit_window.geometry("1000x650")
    audit_window.configure(bg=COLORS['bg'])

    header_frame = tk.Frame(audit_window, bg=COLORS['bg'])
    header_frame.pack(fill='x', padx=20, pady=20)

    title_label = tk.Label(header_frame, text="Audit Log", font=("Inter", 18, "bold"),
                          bg=COLORS['bg'], fg=COLORS['text'])
    title_label.pack(anchor='w')

    content_frame = tk.Frame(audit_window, bg=COLORS['card'], relief='raised', bd=2)
    content_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
    cols = ('ID', 'Action', 'Item ID', 'Item Name', 'User', 'Timestamp')
    tree = ttk.Treeview(content_frame, columns=cols, show='headings', style="Modern.Treeview")
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER, width=120)
    scrollbar_audit = ttk.Scrollbar(content_frame, orient=tk.VERTICAL, command=tree.yview,
                                   style="Modern.Vertical.TScrollbar")
    tree.configure(yscroll=scrollbar_audit.set)

    tree.pack(side=tk.LEFT, fill='both', expand=True, padx=10, pady=10)
    scrollbar_audit.pack(side=tk.RIGHT, fill=tk.Y, pady=10)
    for row in c.execute("SELECT * FROM audit_log ORDER BY timestamp DESC"):
        tree.insert('', 'end', values=row)

def debounce(func, delay):
    def wrapper(*args, **kwargs):
        wrapper.timer = Timer(delay, func, args=args, kwargs=kwargs)
        wrapper.timer.start()
    return wrapper

def on_search(event=None):
    if hasattr(on_search, 'timer'):
        on_search.timer.cancel()
    on_search.debounced_filter()

on_search.debounced_filter = debounce(lambda: display_inventory(search_entry.get()), 0.5)

def generate_excel_report_async():
    def worker():
        timestamp = datetime.now().strftime("%Y_%m_%d")
        default_filename = f"inventory_audit_report_{timestamp}.xlsx"
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename,
                                                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filepath:
            try:
                wb = openpyxl.Workbook()

                # Create inventory sheet
                ws_inventory = wb.active
                ws_inventory.title = "Inventory"
                inventory_headers = ['ID', 'Item Name', 'Quantity', 'Price', 'Updated By']
                ws_inventory.append(inventory_headers)
                for col_num, header in enumerate(inventory_headers, 1):
                    ws_inventory.cell(row=1, column=col_num).font = Font(bold=True)

                c.execute("SELECT id, item_name, quantity, price, updated_by FROM inventory")
                inventory_rows = c.fetchall()

                for row in inventory_rows:
                    ws_inventory.append(row)

                inventory_col_widths = [5, 30, 10, 12, 20]
                for i, width in enumerate(inventory_col_widths, 1):
                    ws_inventory.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

                # Create audit log sheet
                ws_audit = wb.create_sheet(title="Audit Log")
                audit_headers = ['ID', 'Action', 'Item ID', 'Item Name', 'User', 'Timestamp']
                ws_audit.append(audit_headers)
                for col_num, header in enumerate(audit_headers, 1):
                    ws_audit.cell(row=1, column=col_num).font = Font(bold=True)

                c.execute("SELECT id, action, item_id, item_name, user, timestamp FROM audit_log ORDER BY timestamp DESC")
                audit_rows = c.fetchall()

                for row in audit_rows:
                    ws_audit.append(row)

                audit_col_widths = [5, 12, 8, 30, 20, 22]
                for i, width in enumerate(audit_col_widths, 1):
                    ws_audit.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

                wb.save(filepath)
                root.after(0, lambda: messagebox.showinfo("Report Generated", f"Excel report saved successfully:\n{filepath}"))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror("Error", f"Failed to save report: {e}"))

    Thread(target=worker).start()

def show_context_menu(event):
    item = inventory_tree.identify_row(event.y)
    if item:
        inventory_tree.selection_set(item)
        context_menu.post(event.x_root, event.y_root)
    return "break"

def copy_item_details():
    selected = inventory_tree.selection()
    if not selected:
        return

    item_id = selected[0]
    values = inventory_tree.item(item_id)['values']
    if values:
        details = f"ID: {values[0]}\nItem Name: {values[1]}\nQuantity: {values[2]}\nPrice: {values[3]}\nUpdated By: {values[4]}"
        root.clipboard_clear()
        root.clipboard_append(details)
        update_status("Item details copied to clipboard.")

def duplicate_item():
    selected = inventory_tree.selection()
    if not selected:
        return

    item_id = selected[0]
    values = inventory_tree.item(item_id)['values']
    if values:
        entry_widgets["Item Name"].delete(0, tk.END)
        entry_widgets["Item Name"].insert(0, f"Copy of {values[1]}")
        entry_widgets["Quantity"].delete(0, tk.END)
        entry_widgets["Quantity"].insert(0, values[2])
        entry_widgets["Price"].delete(0, tk.END)
        entry_widgets["Price"].insert(0, values[3].replace('$', ''))
        entry_widgets["Updated By"].delete(0, tk.END)
        entry_widgets["Updated By"].insert(0, values[4])
        update_status("Item duplicated. Review and click Add Item to save.")

def treeview_sort_column(tree, col, reverse):
    items = [(tree.set(item, col), item) for item in tree.get_children('')]

    if col in ["Quantity"]:
        items = [(int(value.replace(',', '')), item) for value, item in items]
    elif col in ["Price"]:
        items = [(float(value.replace('$', '').replace(',', '')), item) for value, item in items]
    elif col in ["ID"]:
        items = [(int(value), item) for value, item in items]

    items.sort(reverse=reverse)

    for index, (val, item) in enumerate(items):
        tree.move(item, '', index)

    tree.heading(col, command=lambda: treeview_sort_column(tree, col, not reverse))

def create_modern_button(parent, text, command,
                         bg_color=COLORS['primary'],
                         hover_color=COLORS['primary_hover'],
                         **kwargs):
    kwargs.setdefault("padx", 25)
    kwargs.setdefault("pady", 12)
    button = tk.Button(parent, text=text, command=command,
                       bg=bg_color, fg='white', relief='raised',
                       font=("Inter", 10, "bold"),
                       cursor='hand2', bd=2, highlightthickness=0,
                       **kwargs)
    return button

def create_modern_entry(parent, **kwargs):
    entry = tk.Entry(parent, relief='sunken', bd=2, highlightthickness=1,
                    highlightcolor=COLORS['primary'], highlightbackground=COLORS['border'],
                    font=("Inter", 11), bg='white', fg=COLORS['text'],
                    insertbackground=COLORS['primary'], **kwargs)

    def on_focus_in(e):
        entry.config(highlightcolor=COLORS['primary'], highlightthickness=2, relief='sunken', bd=2)

    def on_focus_out(e):
        entry.config(highlightcolor=COLORS['border'], highlightthickness=1, relief='sunken', bd=1)

    entry.bind('<FocusIn>', on_focus_in)
    entry.bind('<FocusOut>', on_focus_out)

    return entry

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def on_closing():
    if messagebox.askokcancel("Close/End", "Do you want to close the application?"):
        conn.close()
        root.destroy()

# --- Main Application Setup ---
root = tk.Tk()
root.title("Inventory Management System")
root.geometry("1300x950")
root.configure(bg=COLORS['bg'])

icon_path = resource_path("C:/Users/klyde/Documents/Nursing Experts Sevices/Python/Inventory_Project/NE1.ico")
if os.path.exists(icon_path):
    root.iconbitmap(icon_path)
else:
    print(f"Icon file not found at {icon_path}")

style = ttk.Style()
style.theme_use("clam")
style.configure("Modern.Treeview",
                background="white",
                foreground=COLORS['text'],
                rowheight=32,
                fieldbackground="white",
                font=("Inter", 10),
                borderwidth=0)
style.configure("Modern.Treeview.Heading",
                font=("Inter", 11, "bold"),
                background=COLORS['bg'],
                foreground=COLORS['text'])
style.map('Modern.Treeview',
          background=[('selected', COLORS['primary'])],
          foreground=[('selected', 'white')])

style.configure("Modern.Vertical.TScrollbar",
                background=COLORS['bg'],
                troughcolor=COLORS['bg'],
                bordercolor=COLORS['bg'],
                arrowcolor=COLORS['text_light'],
                darkcolor=COLORS['border'],
                lightcolor=COLORS['border'],
                gripcount=0,
                borderwidth=0,
                relief='flat',
                width=12)
style.map("Modern.Vertical.TScrollbar",
          background=[('active', COLORS['border']),
                     ('pressed', COLORS['secondary'])],
          arrowcolor=[('active', COLORS['text']),
                     ('pressed', COLORS['text'])])
style.configure("Modern.Vertical.TScrollbar.thumb",
                background=COLORS['border'],
                relief='flat',
                borderwidth=0)
style.map("Modern.Vertical.TScrollbar.thumb",
          background=[('active', COLORS['secondary']),
                     ('pressed', COLORS['text'])])
# Main frame holds input + inventory table + audit log button
main_frame = tk.Frame(root, bg=COLORS['bg'])
main_frame.pack(fill='both', expand=True, padx=30, pady=20)

# Left side: input and action buttons (fixed width)
input_frame = RoundedFrame(main_frame, bg_color=COLORS['card'], border_color=COLORS['border'],
                           corner_radius=20, border_width=2)
input_frame.pack(side='left', fill='y', padx=(0, 20), pady=10)
input_frame.config(width=440)

# Load and display logo if available
try:
    logo_small_path = resource_path("C:/Users/klyde/Documents/Nursing Experts Sevices/Python/Inventory_Project/NE2.PNG")
    logo_small_img_raw = Image.open(logo_small_path)
    logo_small_img_raw = logo_small_img_raw.resize((300, 60), Image.LANCZOS)
    logo_small_img = ImageTk.PhotoImage(logo_small_img_raw)
    logo_small_label = tk.Label(input_frame.canvas, image=logo_small_img, bg=COLORS['card'])
    logo_small_label.image = logo_small_img
    logo_small_label.place(x=20, y=10)
except Exception as e:
    pass

# Input section title
input_title = tk.Label(input_frame.canvas, text="Item Details", font=("Inter", 18, "bold"),
                       bg=COLORS['card'], fg=COLORS['text'])
input_title.place(x=20, y=120)

# Entry fields labels and entries
labels_texts = ["Item Name", "Quantity", "Price", "Updated By"]
entry_widgets = {}

# Validation commands
vcmd_int = (root.register(validate_non_negative_int), '%P')
vcmd_float = (root.register(validate_non_negative_float), '%P')

for i, label_text in enumerate(labels_texts):
    label = tk.Label(input_frame.canvas, text=label_text, font=("Inter", 12), bg=COLORS['card'], fg=COLORS['text'])
    label.place(x=20, y=170 + i*65)

    validate_cmd = None
    if label_text == "Quantity":
        validate_cmd = vcmd_int
    elif label_text == "Price":
        validate_cmd = vcmd_float
    entry = create_modern_entry(input_frame.canvas, width=30, validate='key', validatecommand=validate_cmd)
    entry.place(x=20, y=195 + i*65)
    entry_widgets[label_text] = entry

# Action Buttons Section
action_frame = tk.Frame(input_frame.canvas, bg=COLORS['card'])
action_frame.place(x=20, y=520, width=380, height=200)
primary_frame = tk.Frame(action_frame, bg=COLORS['card'])
primary_frame.pack(fill='x', pady=(0, 15))
add_button = create_modern_button(primary_frame, "Add Item", add_item, bg_color=COLORS['primary'], padx=20, pady=8)
add_button.pack(side='left', padx=(0, 10))
delete_button = create_modern_button(primary_frame, "Delete Selected", delete_item, bg_color=COLORS['danger'], padx=20, pady=8)
delete_button.pack(side='left')

# Search Section
search_label = tk.Label(input_frame.canvas, text="Search Item:", font=("Inter", 12), bg=COLORS['card'], fg=COLORS['text'])
search_label.place(x=20, y=475)
search_entry = create_modern_entry(input_frame.canvas, width=28)
search_entry.place(x=130, y=475)
search_entry.bind('<KeyRelease>', on_search)

# Audit Section
audit_frame = tk.Frame(action_frame, bg=COLORS['card'])
audit_frame.pack(fill='x', pady=(0, 15))
audit_log_button = create_modern_button(audit_frame, "View Audit Log", show_audit_log, bg_color=COLORS['accent'], padx=20, pady=8)
audit_log_button.pack(side='left', padx=(0, 10))

# Reports Section
reports_frame = tk.Frame(action_frame, bg=COLORS['card'])
reports_frame.pack(fill='x')
report_button = create_modern_button(reports_frame, "Generate Report", generate_excel_report_async, bg_color=COLORS['primary'], padx=30, pady=8)
report_button.pack(side='left')

# Right side: Inventory Table Frame (expanded to fill remaining space)
inventory_frame = RoundedFrame(main_frame, bg_color=COLORS['card'], border_color=COLORS['border'],
                              corner_radius=20, border_width=2)
inventory_frame.pack(side='right', fill='both', expand=True)
inventory_label = tk.Label(inventory_frame.canvas, text="Inventory", font=("Inter", 18, "bold"),
                           bg=COLORS['card'], fg=COLORS['text'])
inventory_label.place(x=0, y=10)

# Treeview for inventory
columns = ("ID", "Item Name", "Quantity", "Price", "Updated By")
inventory_tree = ttk.Treeview(inventory_frame.canvas, columns=columns, show='headings', style="Modern.Treeview")

for col in columns:
    inventory_tree.heading(col, text=col, command=lambda c=col: treeview_sort_column(inventory_tree, c, False))
    if col == "ID":
        inventory_tree.column(col, width=50, minwidth=50, stretch=tk.NO, anchor=tk.CENTER)
    elif col == "Item Name":
        inventory_tree.column(col, width=500, minwidth=200, stretch=tk.NO, anchor=tk.CENTER)
    elif col == "Quantity":
        inventory_tree.column(col, width=70, minwidth=70, stretch=tk.NO, anchor=tk.CENTER)
    elif col == "Price":
        inventory_tree.column(col, width=110, minwidth=80, stretch=tk.NO, anchor=tk.CENTER)
    else:
        inventory_tree.column(col, width=140, anchor=tk.CENTER)

scrollbar = ttk.Scrollbar(inventory_frame.canvas, orient=tk.VERTICAL, command=inventory_tree.yview, style="Modern.Vertical.TScrollbar")
inventory_tree.configure(yscroll=scrollbar.set)
inventory_tree.pack(side='left', fill='both', expand=True, padx=(5, 0), pady=(50, 10))
scrollbar.pack(side='right', fill='y', pady=(60, 10))
inventory_tree.bind('<<TreeviewSelect>>', load_selected_item)
inventory_tree.bind('<Button-3>', show_context_menu)

# Status bar at the bottom
status_var = tk.StringVar()
status_bar = tk.Label(root, textvariable=status_var, bg=COLORS['bg'], fg=COLORS['secondary'], font=("Inter", 12))
status_bar.pack(fill='both', side='bottom', ipady=5)

# Context Menu
context_menu = tk.Menu(root, tearoff=0, bg=COLORS['card'], fg=COLORS['text'],
                      activebackground=COLORS['primary'],
                      activeforeground='white')
context_menu.add_command(label="Edit Item", command=lambda: load_selected_item(None))
context_menu.add_command(label="Delete Item", command=delete_item)
context_menu.add_separator()
context_menu.add_command(label="Duplicate Item", command=duplicate_item)
context_menu.add_command(label="Copy Details", command=copy_item_details)
root.option_add('*Menu.borderWidth', '1')
root.option_add('*Menu.activeBorderWidth', '1')
root.option_add('*Menu.relief', 'solid')
root.option_add('*Menu.font', ('Inter', 10))

# Initialize database connection and setup
conn = get_db_connection()
c = conn.cursor()
initialize_database()

# Initial load of inventory list
display_inventory()
update_status("Ready.")

root.protocol("WM_DELETE_WINDOW", on_closing)
center_window(root, 1366, 900)

root.mainloop()
