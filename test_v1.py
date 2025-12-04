
# pip install pyinstaller

import tkinter as tk
from PIL import Image, ImageTk
import tkinter.font as tkfont
from tkinter import messagebox, ttk, filedialog
import sqlite3
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import sys
import os



# --- Database Setup ---
try:
    conn = sqlite3.connect('inventory.db')
    c = conn.cursor()

    # Create inventory table if not exists
    c.execute('''
    CREATE TABLE IF NOT EXISTS inventory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_name TEXT NOT NULL,
        quantity INTEGER NOT NULL CHECK(quantity >= 0),
        price REAL NOT NULL CHECK(price >= 0),
        updated_by TEXT NOT NULL
    )
    ''')

    # Create audit log table if not exists
    c.execute('''
    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action TEXT NOT NULL,
        item_id INTEGER,
        item_name TEXT,
        user TEXT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    conn.commit()
except sqlite3.Error as e:
    messagebox.showerror("Database Error", f"Failed to initialize database: {e}")
    exit(1)

selected_item_id = None  # Track item being edited


# --- Input Validation Functions ---
def validate_non_negative_int(P):
    if P == "":
        return True
    if P.isdigit() and int(P) >= 0:
        return True
    return False


def validate_non_negative_float(P):
    if P == "":
        return True
    try:
        value = float(P)
        return value >= 0
    except ValueError:
        return False


# --- Functions ---
def update_status(message):
    status_var.set(message)


def clear_inputs():
    global selected_item_id
    selected_item_id = None
    entry_widgets["Item Name"].delete(0, tk.END)
    entry_widgets["Quantity"].delete(0, tk.END)
    entry_widgets["Price"].delete(0, tk.END)
    entry_widgets["Updated By"].delete(0, tk.END)
    add_button.config(text="Add Item")
    update_status("Ready to add new item.")


def display_inventory(filter_text=""):
    for row in inventory_tree.get_children():
        inventory_tree.delete(row)
    query = "SELECT * FROM inventory"
    params = ()
    if filter_text:
        query += " WHERE item_name LIKE ?"
        params = ('%' + filter_text + '%',)
    for row in c.execute(query, params):
        inventory_tree.insert('', 'end', iid=row[0], values=(row[0], row[1], row[2], f"${row[3]:.2f}", row[4] if row[4] else 'N/A'))


def load_selected_item(event):
    global selected_item_id
    selected = inventory_tree.selection()
    if selected:
        selected_item_id = int(selected[0])
        c.execute("SELECT * FROM inventory WHERE id=?", (selected_item_id,))
        row = c.fetchone()
        if row:
            entry_widgets["Item Name"].delete(0, tk.END)
            entry_widgets["Item Name"].insert(0, row[1])
            entry_widgets["Quantity"].delete(0, tk.END)
            entry_widgets["Quantity"].insert(0, row[2])
            entry_widgets["Price"].delete(0, tk.END)
            entry_widgets["Price"].insert(0, row[3])
            entry_widgets["Updated By"].delete(0, tk.END)
            entry_widgets["Updated By"].insert(0, row[4] if row[4] else '')
            add_button.config(text="Update Item")
            update_status(f"Loaded item ID {selected_item_id} for editing.")
    else:
        clear_inputs()
def is_duplicate_name(name, exclude_id=None):
    if exclude_id is None:
        c.execute("SELECT COUNT(*) FROM inventory WHERE LOWER(item_name) = LOWER(?)", (name,))
    else:
        c.execute("SELECT COUNT(*) FROM inventory WHERE LOWER(item_name) = LOWER(?) AND id != ?", (name, exclude_id))
    count = c.fetchone()[0]
    return count > 0

def add_item():
    global selected_item_id
    name = entry_widgets["Item Name"].get().strip()
    quantity = entry_widgets["Quantity"].get().strip()
    price = entry_widgets["Price"].get().strip()
    user = entry_widgets["Updated By"].get().strip()

    if not (name and quantity and price and user):
        update_status("All fields are required!")
        return

    # Validate numeric fields
    try:
        quantity_int = int(quantity)
        price_float = float(price)
        if quantity_int < 0 or price_float < 0:
            update_status("Quantity and Price must be non-negative.")
            return
    except ValueError:
        update_status("Quantity must be an integer and Price must be a number.")
        return

    # Check for duplicate names (exclude current item if updating)
    if is_duplicate_name(name, exclude_id=selected_item_id):
        update_status(f"Item name '{name}' already exists.")
        return

    try:
        if selected_item_id is None:
            # Insert new item
            c.execute(
                "INSERT INTO inventory (item_name, quantity, price, updated_by) VALUES (?, ?, ?, ?)",
                (name, quantity_int, price_float, user)
            )
            item_id = c.lastrowid
            action = "Added"
        else:
            # Update existing item
            c.execute(
                "UPDATE inventory SET item_name=?, quantity=?, price=?, updated_by=? WHERE id=?",
                (name, quantity_int, price_float, user, selected_item_id)
            )
            item_id = selected_item_id
            action = "Updated"

        # Log the action
        c.execute(
            "INSERT INTO audit_log (action, item_id, item_name, user) VALUES (?, ?, ?, ?)",
            (action, item_id, name, user)
        )
        conn.commit()

        messagebox.showinfo("Success", f"Item {action.lower()} successfully!")
        clear_inputs()
        display_inventory()
        update_status(f"Item {action.lower()} successfully.")
    except Exception as e:
        messagebox.showerror("Error", str(e))
        


def delete_item():
    selected = inventory_tree.selection()
    if selected:
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete the selected item?"):
            update_status("Delete cancelled.")
            return
        item_id = int(selected[0])
        c.execute("SELECT item_name, updated_by FROM inventory WHERE id=?", (item_id,))
        row = c.fetchone()
        if row:
            item_name, user = row
        else:
            item_name, user = "Unknown", "Unknown"
        c.execute("DELETE FROM inventory WHERE id=?", (item_id,))
        c.execute("INSERT INTO audit_log (action, item_id, item_name, user) VALUES (?, ?, ?, ?)",
                  ("Deleted", item_id, item_name, user))
        conn.commit()
        messagebox.showinfo("Success", "Item deleted successfully!")
        display_inventory()
        clear_inputs()
        update_status("Item deleted successfully.")
    else:
        update_status("Select an item to delete.")


def show_audit_log():
    audit_window = tk.Toplevel(root)
    audit_window.title("Audit Log")
    audit_window.geometry("900x600")

    cols = ('ID', 'Action', 'Item ID', 'Item Name', 'User', 'Timestamp')
    tree = ttk.Treeview(audit_window, columns=cols, show='headings')

    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER, width=110)

    tree.pack(fill='both', expand=True)

    for row in c.execute("SELECT * FROM audit_log ORDER BY timestamp DESC"):
        tree.insert('', 'end', values=row)


def generate_excel_report():
    timestamp = datetime.now().strftime("%Y_%m_%d")
    default_filename = f"inventory_audit_report_{timestamp}.xlsx"
    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename,
                                            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if not filepath:
        return  # Cancelled

    wb = openpyxl.Workbook()

    # --- Inventory Sheet ---
    ws_inventory = wb.active
    ws_inventory.title = "Inventory"

    inventory_headers = ['ID', 'Item Name', 'Quantity', 'Price', 'Updated By']
    ws_inventory.append(inventory_headers)

    for col_num, header in enumerate(inventory_headers, 1):
        ws_inventory.cell(row=1, column=col_num).font = Font(bold=True)

    c.execute("SELECT id, item_name, quantity, price, updated_by FROM inventory")
    inventory_rows = c.fetchall()
    for row in inventory_rows:
        ws_inventory.append(row)

    inventory_col_widths = [5, 30, 10, 12, 20]
    for i, width in enumerate(inventory_col_widths, 1):
        ws_inventory.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # --- Audit Log Sheet ---
    ws_audit = wb.create_sheet(title="Audit Log")

    audit_headers = ['ID', 'Action', 'Item ID', 'Item Name', 'User', 'Timestamp']
    ws_audit.append(audit_headers)

    for col_num, header in enumerate(audit_headers, 1):
        ws_audit.cell(row=1, column=col_num).font = Font(bold=True)

    c.execute("SELECT id, action, item_id, item_name, user, timestamp FROM audit_log ORDER BY timestamp DESC")
    audit_rows = c.fetchall()
    for row in audit_rows:
        ws_audit.append(row)

    audit_col_widths = [5, 12, 8, 30, 20, 22]
    for i, width in enumerate(audit_col_widths, 1):
        ws_audit.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    try:
        wb.save(filepath)
        messagebox.showinfo("Report Generated", f"Excel report saved successfully:\n{filepath}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save report: {e}")



def on_search(event):
    filter_text = search_entry.get()
    display_inventory(filter_text)


def new_item():
    clear_inputs()

#IMAGE PATH
def resource_path(relative_path):
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

logo_path = resource_path("NE2.PNG")  
print(f"Image path: {logo_path}")


# --- GUI Setup ---
root = tk.Tk()
root.title("Inventory Monitoring System")
root.geometry("1200x900")
root.configure(bg="#F7F7F7")  #  background color

# Load and display logo image


logo_img_path = resource_path("Inventory_Project/NE2.PNG")  
logo_img_raw = Image.open(logo_img_path)
logo_img_raw = logo_img_raw.resize((600, 60), Image.LANCZOS)
logo_img = ImageTk.PhotoImage(logo_img_raw)

logo_frame = tk.Frame(root, bg="#f4f4f4")
logo_frame.pack(padx=10, pady=(10, 0), anchor='center')

logo_label = tk.Label(logo_frame, image=logo_img, bg="#f4f4f4")
logo_label.pack()

style = ttk.Style()
style.theme_use("clam")

# Treeview styling
style.configure("Treeview",
                background="white",
                foreground="black",
                rowheight=25,
                fieldbackground="white",
                font=("Segoe UI", 10))
style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"))
style.map('Treeview', background=[('selected', '#cce5ff')])

# Fonts
label_font = ("Segoe UI", 10)
entry_font = ("Segoe UI", 10)

# Search + Action Bar
search_frame = tk.Frame(root, bg="#f4f4f4")
search_frame.pack(pady=10, padx=10, fill='x')

tk.Label(search_frame, text="Search:", font=label_font, bg="#f4f4f4").pack(side=tk.LEFT)
search_entry = tk.Entry(search_frame, font=entry_font, width=25)
search_entry.pack(side=tk.LEFT, padx=(5, 15))
search_entry.bind("<KeyRelease>", on_search)



btn_style = {'font': label_font, 'bg': "#976730", 'fg': 'white', 'activebackground': '#005f99', 'width': 16}

new_item_button = tk.Button(search_frame, text="New Item", command=new_item, **btn_style)
new_item_button.pack(side=tk.LEFT, padx=5)


generate_excel_button = tk.Button(search_frame, text="Generate Report", command=generate_excel_report, **btn_style)
generate_excel_button.pack(side=tk.LEFT, padx=5)


audit_button = tk.Button(search_frame, text="View Audit Log", command=show_audit_log, **btn_style)
audit_button.pack(side=tk.LEFT, padx=5)

# Input Section
input_frame = tk.Frame(root, bg="#f4f4f4")
input_frame.pack(pady=10)

entries = [
    ("Item Name", 20),
    ("Quantity", 8),
    ("Price", 10),
    ("Updated By", 15)
]
entry_widgets = {}

for i, (label, width) in enumerate(entries):
    tk.Label(input_frame, text=label, font=label_font, bg="#f4f4f4").grid(row=0, column=2 * i, padx=5, sticky='e')
    entry = tk.Entry(input_frame, width=width, font=entry_font)
    entry.grid(row=0, column=2 * i + 1, padx=5, pady=3)
    entry_widgets[label] = entry

add_button = tk.Button(input_frame, text="Add Item", command=add_item, **btn_style)
add_button.grid(row=1, column=0, columnspan=2, pady=10)

delete_button = tk.Button(input_frame, text="Delete Selected", command=delete_item, **btn_style)
delete_button.grid(row=1, column=2, columnspan=2, pady=10)

# Inventory List (Treeview)
list_frame = tk.Frame(root, padx=10, pady=5)
list_frame.pack(fill='both', expand=True)

columns = ('ID', 'Name', 'Qty', 'Price', 'Updated By')
inventory_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)

column_widths = {'ID': 30, 'Name': 180, 'Qty': 50, 'Price': 70, 'Updated By': 100}
for col in columns:
    inventory_tree.heading(col, text=col)
    inventory_tree.column(col, anchor=tk.CENTER, width=column_widths[col])

inventory_tree.pack(side=tk.LEFT, fill='both', expand=True)

# Scrollbar for Treeview
scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=inventory_tree.yview)
inventory_tree.configure(yscroll=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

inventory_tree.bind('<<TreeviewSelect>>', load_selected_item)

display_inventory()


#display bottom status bar
status_var = tk.StringVar()
status_var.set("Ready.")

# Create Frame with fixed height for status bar
status_frame = tk.Frame(root, height=25, bg='white')
status_frame.pack(side=tk.BOTTOM, fill=tk.X)
status_frame.pack_propagate(False)  # Prevent shrinking

# Create Label inside the Frame filling it
status_bar = tk.Label(
    status_frame,
    textvariable=status_var,
    relief=tk.SUNKEN,
    anchor='w',
    bg='white',
    font=("Segoe UI", 12)
)
status_bar.pack(fill=tk.BOTH, expand=True)



# Exit handler to close DB connection cleanly
def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        conn.close()
        root.destroy()


root.protocol("WM_DELETE_WINDOW", on_closing)

# Input validation setup
vcmd_int = (root.register(validate_non_negative_int), '%P')
vcmd_float = (root.register(validate_non_negative_float), '%P')

entry_widgets["Quantity"].config(validate='key', validatecommand=vcmd_int)
entry_widgets["Price"].config(validate='key', validatecommand=vcmd_float)

root.mainloop()


#python -m PyInstaller